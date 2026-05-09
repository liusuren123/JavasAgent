"""Excel 表格操作子模块。

提供 Excel (.xlsx) 文件的创建、读取、追加、编辑能力。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from loguru import logger

from src.utils.path_safety import PathSafetyError, safe_resolve_path


# ======================================================================
# 公共辅助
# ======================================================================


def _safe_path(workspace: Path, user_path: str, *, allow_create_parents: bool = False) -> Path:
    """安全解析用户路径。"""
    return safe_resolve_path(workspace, user_path, allow_create_parents=allow_create_parents)


def _check_openpyxl_import() -> str | None:
    """检查 openpyxl 是否已安装。"""
    try:
        import openpyxl  # noqa: F401
        return None
    except ImportError:
        return "openpyxl 未安装，请运行: pip install openpyxl"


# ======================================================================
# 操作实现
# ======================================================================


async def read_xlsx(workspace: Path, params: dict) -> dict:
    """读取 Excel 文件内容。

    Params:
        path: 文件路径
        sheet: 工作表名称（默认第一个）
        max_rows: 最大行数（默认 1000）
        include_formulas: 是否包含公式（默认 False）
    """
    try:
        path = _safe_path(workspace, params["path"])
    except PathSafetyError as e:
        return {"error": str(e)}

    if not path.exists():
        return {"error": f"文件不存在: {path}"}

    err = _check_openpyxl_import()
    if err:
        return {"error": err}

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), data_only=not params.get("include_formulas", False))
        sheet_name = params.get("sheet")
        ws = wb[sheet_name] if sheet_name else wb.active

        max_rows = params.get("max_rows", 1000)
        rows: list[list[Any]] = []
        for row in ws.iter_rows(max_row=min(ws.max_row, max_rows), values_only=True):
            rows.append(list(row))

        sheet_names = wb.sheetnames
        wb.close()

        logger.info(f"读取 Excel: {path} ({ws.title}, {len(rows)} 行)")
        return {
            "path": str(path),
            "sheet": ws.title,
            "sheet_names": sheet_names,
            "rows": rows,
            "row_count": len(rows),
            "column_count": ws.max_column,
        }
    except KeyError:
        return {"error": f"工作表不存在: {sheet_name}，可用: {wb.sheetnames}"}
    except Exception as e:
        logger.error(f"读取 Excel 失败: {e}")
        return {"error": f"读取失败: {e}"}


async def create_xlsx(workspace: Path, params: dict) -> dict:
    """创建 Excel 文件。

    Params:
        path: 文件路径
        sheet: 工作表名称（默认 "Sheet1"）
        headers: 列标题列表
        rows: 数据行列表（二维数组）
        column_widths: 列宽列表（可选）
    """
    try:
        path = _safe_path(workspace, params["path"], allow_create_parents=True)
    except PathSafetyError as e:
        return {"error": str(e)}

    err = _check_openpyxl_import()
    if err:
        return {"error": err}

    try:
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        sheet_name = params.get("sheet", "Sheet1")
        ws = wb.active
        ws.title = sheet_name

        headers = params.get("headers", [])
        rows = params.get("rows", [])

        # 写入表头
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # 写入数据行
        for row in rows:
            ws.append(row)

        # 设置列宽
        widths = params.get("column_widths")
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        elif headers:
            for i, h in enumerate(headers, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(
                    len(str(h)) * 2 + 4, 12
                )

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        wb.close()

        total_rows = len(rows) + (1 if headers else 0)
        logger.info(f"创建 Excel: {path} ({total_rows} 行, {len(headers)} 列)")
        return {
            "path": str(path),
            "created": True,
            "sheet": sheet_name,
            "rows": total_rows,
            "columns": len(headers),
        }
    except Exception as e:
        logger.error(f"创建 Excel 失败: {e}")
        return {"error": f"创建失败: {e}"}


async def append_xlsx(workspace: Path, params: dict) -> dict:
    """追加数据到现有 Excel 文件。

    Params:
        path: 文件路径
        rows: 要追加的数据行列表
        sheet: 工作表名称（默认活动工作表）
    """
    try:
        path = _safe_path(workspace, params["path"])
    except PathSafetyError as e:
        return {"error": str(e)}

    if not path.exists():
        return {"error": f"文件不存在: {path}"}

    err = _check_openpyxl_import()
    if err:
        return {"error": err}

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path))
        sheet_name = params.get("sheet")
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = params.get("rows", [])
        for row in rows:
            ws.append(row)

        wb.save(str(path))
        wb.close()

        logger.info(f"追加 Excel: {path} (+{len(rows)} 行)")
        return {"path": str(path), "appended_rows": len(rows)}
    except Exception as e:
        logger.error(f"追加 Excel 失败: {e}")
        return {"error": f"追加失败: {e}"}


async def edit_xlsx(workspace: Path, params: dict) -> dict:
    """编辑已有 Excel 文件。

    支持：
    1. 修改单元格值：指定 ``cells`` 列表，每项含 row / col / value
    2. 添加行：指定 ``add_rows`` 列表，每项为数据数组
    3. 删除行：指定 ``delete_rows`` 列表，每项为行号（1-based）

    Params:
        path: 文件路径
        sheet: 工作表名称（默认活动工作表）
        cells: 单元格修改列表 [{"row": 2, "col": 1, "value": "新值"}]
        add_rows: 追加行列表 [[值1, 值2, ...], ...]
        delete_rows: 删除行号列表 [3, 5]（1-based，按从大到小排序删除）
    """
    try:
        path = _safe_path(workspace, params["path"])
    except PathSafetyError as e:
        return {"error": str(e)}

    if not path.exists():
        return {"error": f"文件不存在: {path}"}

    err = _check_openpyxl_import()
    if err:
        return {"error": err}

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path))
        sheet_name = params.get("sheet")
        ws = wb[sheet_name] if sheet_name else wb.active

        changes = 0

        # 修改单元格值
        for cell_spec in params.get("cells", []):
            row_num = cell_spec["row"]
            col_num = cell_spec["col"]
            value = cell_spec["value"]
            ws.cell(row=row_num, column=col_num, value=value)
            changes += 1

        # 删除行（从大到小删除，避免索引偏移）
        delete_rows = sorted(params.get("delete_rows", []), reverse=True)
        for row_num in delete_rows:
            ws.delete_rows(row_num)
            changes += 1

        # 添加行
        for row_data in params.get("add_rows", []):
            ws.append(row_data)
            changes += 1

        wb.save(str(path))
        wb.close()

        logger.info(f"编辑 Excel: {path} ({changes} 处修改)")
        return {"path": str(path), "changes": changes}
    except Exception as e:
        logger.error(f"编辑 Excel 失败: {e}")
        return {"error": f"编辑失败: {e}"}
